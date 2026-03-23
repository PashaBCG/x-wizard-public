"""
scan_excel.py — Lightweight Excel file scanner for X-Wizard skill creation.

Usage:
    python scan_excel.py <file_path> [--rows 20] [--sheet <sheet_name>]

Output:
    JSON to stdout with the following structure:
    {
        "file": "<path>",
        "file_size_mb": 12.4,
        "sheets": ["Data", "Summary", "Cover"],
        "selected_sheet": "Data",
        "row_count": 523,
        "columns": [
            {
                "name": "Office",
                "sample_values": ["London", "Boston", "NYC"],
                "non_null_count": 521
            },
            ...
        ]
    }

Exit codes:
    0 = success
    1 = file not found
    2 = unsupported file type
    3 = read error (file locked, corrupt, etc.)

This script is used as a fallback when the Excel MCP is unavailable.
It is shared across the create-skill and start skills.
Reference: .cursor/skills/create-skill/scripts/scan_excel.py
"""

import sys
import json
import os
import argparse


def find_best_data_sheet(wb):
    """
    Identify the most likely raw-data sheet: the one with the most rows
    that isn't obviously a summary, chart, or cover sheet.
    """
    skip_keywords = {"summary", "cover", "chart", "index", "toc", "readme", "instructions"}
    best_sheet = None
    best_row_count = 0

    for sheet_name in wb.sheetnames:
        lower = sheet_name.lower()
        if any(kw in lower for kw in skip_keywords):
            continue
        ws = wb[sheet_name]
        row_count = ws.max_row
        if row_count > best_row_count:
            best_row_count = row_count
            best_sheet = sheet_name

    # If all sheets were filtered out, fall back to the first one
    if best_sheet is None and wb.sheetnames:
        best_sheet = wb.sheetnames[0]

    return best_sheet


def scan_sheet(ws, max_rows=20):
    """
    Read up to max_rows rows from a worksheet.
    Returns headers and up to 3 unique sample values per column.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], 0, []

    # Find the header row: first row with mostly non-None values
    header_row_idx = 0
    for i, row in enumerate(rows[:5]):
        non_null = sum(1 for cell in row if cell is not None)
        if non_null >= max(1, len(row) // 2):
            header_row_idx = i
            break

    headers = [str(h) if h is not None else f"Col_{i}" for i, h in enumerate(rows[header_row_idx])]
    data_rows = rows[header_row_idx + 1 : header_row_idx + 1 + max_rows]
    total_data_rows = ws.max_row - header_row_idx - 1

    columns = []
    for col_idx, col_name in enumerate(headers):
        values = []
        seen = set()
        non_null_count = 0
        for row in data_rows:
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None:
                non_null_count += 1
                str_val = str(val).strip()
                if str_val and str_val not in seen and len(values) < 3:
                    values.append(str_val)
                    seen.add(str_val)

        columns.append({
            "name": col_name,
            "sample_values": values,
            "non_null_count": non_null_count
        })

    return headers, total_data_rows, columns


def main():
    parser = argparse.ArgumentParser(description="Scan an Excel file for column structure and sample values.")
    parser.add_argument("file_path", help="Path to the .xlsx file")
    parser.add_argument("--rows", type=int, default=20, help="Number of data rows to sample (default: 20)")
    parser.add_argument("--sheet", type=str, default=None, help="Sheet name to scan (default: auto-detect)")
    args = parser.parse_args()

    file_path = args.file_path

    if not os.path.exists(file_path):
        print(json.dumps({"error": f"File not found: {file_path}"}), file=sys.stderr)
        sys.exit(1)

    _, ext = os.path.splitext(file_path.lower())
    if ext not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        print(json.dumps({"error": f"Unsupported file type: {ext}. Expected .xlsx or similar."}), file=sys.stderr)
        sys.exit(2)

    try:
        import openpyxl
    except ImportError:
        print(json.dumps({"error": "openpyxl is not installed. Run: pip install openpyxl"}), file=sys.stderr)
        sys.exit(3)

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        print(json.dumps({"error": f"Could not open file: {e}"}), file=sys.stderr)
        sys.exit(3)

    file_size_mb = round(os.path.getsize(file_path) / (1024 * 1024), 2)
    sheet_names = wb.sheetnames

    selected_sheet = args.sheet if args.sheet else find_best_data_sheet(wb)

    if selected_sheet not in wb.sheetnames:
        print(json.dumps({"error": f"Sheet '{selected_sheet}' not found. Available: {sheet_names}"}), file=sys.stderr)
        sys.exit(3)

    ws = wb[selected_sheet]
    headers, row_count, columns = scan_sheet(ws, max_rows=args.rows)

    result = {
        "file": file_path,
        "file_size_mb": file_size_mb,
        "sheets": sheet_names,
        "selected_sheet": selected_sheet,
        "row_count": row_count,
        "columns": columns
    }

    print(json.dumps(result, indent=2, ensure_ascii=False))
    wb.close()
    sys.exit(0)


if __name__ == "__main__":
    main()
